"""Testes para ``excel_exporter`` (Fase 3 / 3.3)."""
from __future__ import annotations

import sys
import tempfile
import time
import unittest
from pathlib import Path
from unittest import mock

from Virus_project import ScanResult
from excel_exporter import ExcelReportGenerator
from report_generator import ReportMetadata


def _results():
    return [
        ScanResult(file_path="/tmp/a.txt", status="clean", sha256="a" * 64),
        ScanResult(file_path="/tmp/b.txt", status="clean", sha256="b" * 64),
        ScanResult(
            file_path="/tmp/bad/evil.exe",
            status="infected",
            reason="Trojan.A",
            sha256="c" * 64,
        ),
        ScanResult(
            file_path="/tmp/bad/worm.bin",
            status="infected",
            reason="Worm.B",
            sha256="d" * 64,
        ),
        ScanResult(file_path="/tmp/skip", status="skip"),
    ]


class TestExcelExporter(unittest.TestCase):
    def test_is_available_true_when_openpyxl_installed(self):
        self.assertTrue(ExcelReportGenerator.is_available())

    def test_generate_returns_false_when_openpyxl_missing(self):
        with mock.patch.dict(sys.modules, {"openpyxl": None}):
            with tempfile.TemporaryDirectory() as tmp:
                out = Path(tmp) / "r.xlsx"
                ok = ExcelReportGenerator.generate(_results(), out)
                self.assertFalse(ok)

    def test_generate_creates_expected_sheets(self):
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "r.xlsx"
            meta = ReportMetadata(
                scan_id="abc",
                started_at=time.time() - 4,
                finished_at=time.time(),
                paths=["/tmp"],
            )
            self.assertTrue(ExcelReportGenerator.generate(_results(), out, metadata=meta))
            wb = load_workbook(out)
            self.assertEqual(wb.sheetnames, ["Summary", "Infected", "Clean", "Stats"])

            summary_values = {
                wb["Summary"].cell(row=r, column=1).value: wb["Summary"].cell(row=r, column=2).value
                for r in range(2, wb["Summary"].max_row + 1)
            }
            self.assertEqual(summary_values["Total analisado"], 5)
            self.assertEqual(summary_values["Limpos"], 2)
            self.assertEqual(summary_values["Infectados"], 2)
            self.assertEqual(summary_values["Ignorados"], 1)
            self.assertEqual(summary_values["Scan ID"], "abc")

            self.assertEqual(wb["Infected"].max_row, 3)
            self.assertEqual(wb["Infected"].cell(row=2, column=2).value, "Trojan.A")

    def test_clean_sheet_caps_rows(self):
        from openpyxl import load_workbook

        many_clean = [
            ScanResult(file_path=f"/x/{i}", status="clean", sha256="a" * 64)
            for i in range(150)
        ]
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "r.xlsx"
            self.assertTrue(
                ExcelReportGenerator.generate(many_clean, out, max_clean_rows=50)
            )
            wb = load_workbook(out)
            ws = wb["Clean"]
            self.assertEqual(ws.max_row, 1 + 50 + 1)
            self.assertIn("omitidos", str(ws.cell(row=ws.max_row, column=1).value))

    def test_no_infected_uses_placeholder_row(self):
        from openpyxl import load_workbook

        clean_only = [ScanResult(file_path="/x", status="clean", sha256="a" * 64)]
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp) / "r.xlsx"
            self.assertTrue(ExcelReportGenerator.generate(clean_only, out))
            wb = load_workbook(out)
            ws = wb["Infected"]
            self.assertIn("sem ameaças", str(ws.cell(row=2, column=1).value))


if __name__ == "__main__":
    unittest.main(verbosity=2)
