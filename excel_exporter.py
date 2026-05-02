"""
Exportação de relatórios para Excel (.xlsx) — Fase 3 (3.3) do plano.

Gera quatro sheets:

* **Summary** — métricas agregadas (total/clean/infected/skipped, scan rate, paths).
* **Infected** — todas as ameaças detectadas com hash e razão.
* **Clean** — ficheiros limpos (limitado a ``max_clean_rows`` por defeito).
* **Stats** — top diretórios afectados e distribuição por estado.

A dependência ``openpyxl`` é importada lazy: se ausente, ``ExcelReportGenerator.generate``
devolve ``False`` e regista o erro, sem rebentar o resto do app.
"""
from __future__ import annotations

import logging
from collections import Counter
from pathlib import Path
from typing import List, Optional, Sequence

from Virus_project import ScanResult
from report_generator import ReportMetadata, _format_duration, _top_infected_dirs

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """Gera relatórios .xlsx multi-sheet a partir de ``ScanResult``s."""

    @staticmethod
    def is_available() -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    @classmethod
    def generate(
        cls,
        results: Sequence[ScanResult],
        output_file: Path = Path("output/scan_report.xlsx"),
        metadata: Optional[ReportMetadata] = None,
        max_clean_rows: int = 1000,
    ) -> bool:
        """Gera ficheiro Excel. Retorna ``False`` se ``openpyxl`` ausente ou erro I/O."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error(
                "openpyxl não instalado — `pip install openpyxl` para activar exportação Excel."
            )
            return False

        try:
            output_file.parent.mkdir(parents=True, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="667EEA")
            infected_fill = PatternFill("solid", fgColor="FFCDD2")
            center = Alignment(horizontal="center", vertical="center")

            def _style_header(ws, row: int, columns: int) -> None:
                for col in range(1, columns + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center

            def _autosize(ws, max_width: int = 80) -> None:
                for col_idx, col_cells in enumerate(ws.columns, start=1):
                    width = 8
                    for cell in col_cells:
                        if cell.value is None:
                            continue
                        width = max(width, min(max_width, len(str(cell.value)) + 2))
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

            cls._build_summary_sheet(
                wb,
                results,
                metadata,
                _style_header,
                _autosize,
            )
            cls._build_infected_sheet(
                wb,
                results,
                _style_header,
                _autosize,
                infected_fill,
            )
            cls._build_clean_sheet(
                wb,
                results,
                max_clean_rows,
                _style_header,
                _autosize,
            )
            cls._build_stats_sheet(
                wb,
                results,
                _style_header,
                _autosize,
            )

            wb.save(str(output_file))
            logger.info(f"Relatório Excel gerado em {output_file}")
            return True
        except (OSError, IOError, ValueError) as e:
            logger.error(f"Erro ao gerar relatório Excel: {e}")
            return False

    @staticmethod
    def _build_summary_sheet(wb, results, metadata, style_header, autosize):
        ws = wb.create_sheet("Summary")
        counts = Counter(r.status for r in results)

        rows = [
            ("Métrica", "Valor"),
            ("Total analisado", len(results)),
            ("Limpos", counts.get("clean", 0)),
            ("Infectados", counts.get("infected", 0)),
            ("Ignorados", counts.get("skip", 0)),
        ]
        if metadata is not None:
            rows.extend(
                [
                    ("Scan ID", metadata.scan_id),
                    ("Início", metadata.started_iso),
                    ("Duração", _format_duration(metadata.duration_seconds)),
                    (
                        "Taxa",
                        f"{metadata.scan_rate(len(results)):.2f} ficheiros/s",
                    ),
                    ("Pastas", " · ".join(metadata.paths) if metadata.paths else "—"),
                ]
            )
        for row in rows:
            ws.append(row)
        style_header(ws, 1, 2)
        autosize(ws)

    @staticmethod
    def _build_infected_sheet(wb, results, style_header, autosize, infected_fill):
        ws = wb.create_sheet("Infected")
        ws.append(["Ficheiro", "Ameaça", "SHA256"])
        style_header(ws, 1, 3)
        infected = [r for r in results if r.status == "infected"]
        for r in infected:
            ws.append([r.file_path, r.reason or "(desconhecida)", r.sha256])
            for col in range(1, 4):
                ws.cell(row=ws.max_row, column=col).fill = infected_fill
        if not infected:
            ws.append(["(sem ameaças detectadas)", "", ""])
        autosize(ws)

    @staticmethod
    def _build_clean_sheet(wb, results, max_rows, style_header, autosize):
        ws = wb.create_sheet("Clean")
        ws.append(["Ficheiro", "SHA256"])
        style_header(ws, 1, 2)
        clean = [r for r in results if r.status == "clean"]
        for r in clean[:max_rows]:
            ws.append([r.file_path, r.sha256])
        if len(clean) > max_rows:
            ws.append([f"... + {len(clean) - max_rows} ficheiros omitidos", ""])
        autosize(ws)

    @staticmethod
    def _build_stats_sheet(wb, results, style_header, autosize):
        ws = wb.create_sheet("Stats")
        counts = Counter(r.status for r in results)
        ws.append(["Distribuição por estado", "Total"])
        style_header(ws, 1, 2)
        for status in ("clean", "infected", "skip"):
            ws.append([status, counts.get(status, 0)])

        ws.append([])
        header_row = ws.max_row + 1
        ws.append(["Top diretórios com ameaças", "Total"])
        style_header(ws, header_row, 2)
        top = _top_infected_dirs(results, limit=20)
        if top:
            for directory, count in top:
                ws.append([directory, count])
        else:
            ws.append(["(sem ameaças)", 0])
        autosize(ws)
